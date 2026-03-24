import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.colors import Color
from datetime import date, timedelta
import math

HOURLY_RATE = 100
PACE = 2
MAX_C_LEN = 72
CONSULTATION_NUMBER = 149
TIERS = ["Simple", "Base", "Premium"]
BASE_TIER = "Base"
CONSULTATIONS_DIR = r"C:\Users\twpot\Documents\Tyler's Stuff\Tyler's Documents\Investing\Freelancing\Consultations"
output_path = f"{CONSULTATIONS_DIR}\\Consultation{CONSULTATION_NUMBER}.xlsx"

# --- Line item data ---
# Tier key: S=Simple($2k), B=Base($3k), P=Premium($4k)
# Scalar = same hours all tiers. Dict = per-tier hours.
SBP = {"Simple": 0, "Base": 0.5, "Premium": 0.5}  # Base + Premium
P_ONLY = {"Simple": 0, "Base": 0, "Premium": 0.5}  # Premium only

items = [
    # === PHASE 1: Preparation ===
    ("phase", "Preparation"),
    ("item", "Reading/understanding requirements", 0.5),
    ("item", "Interaction with the client throughout", 0.5),
    ("item", "Heroku cloud application setup", 0.5),
    ("item", "MongoDB cloud database setup", 0.5),
    ("item", "Gmail email service setup", SBP),
    ("item", "Twilio texting service setup (not using)", 0),
    ("item", "Discord notification service setup (not using)", 0),
    ("item", "Login system", 0.5),

    # === PHASE 2: Programming Back-End ===
    ("phase", "Programming Back-End "),
    ("category", "Copy trading engine"),
    ("item", "Retrieve settings from database", 0.5),
    ("item", "Poll master account for new orders via API", 0.5),
    ("item", "Detect new orders, filter already-processed orders", 0.5),
    ("item", "Reconstruct multi-leg order from master order data", 0.5),
    ("item", "Apply position size multiplier per follower account", 0.5),
    ("item", "Forward reconstructed order to all follower accounts", 0.5),
    ("item", "Detect master cancellations, cancel on followers", 0.5),
    ("item", "Stale order timeout and auto-cancel logic", 0.5),
    ("item", "Store processed order IDs to prevent duplicates", 0.5),
    ("item", "Monitor fill status of all orders, record when filled", 0.5),
    ("item", "Market hours and early close detection", 0.5),
    ("item", "Logging functions to track what happens at every step", 0.5),
    ("item", "Implement email notifications for order fills", SBP),
    ("item", "Trade direction control (long/short filtering)", SBP),
    ("item", "Trade reversal (mirror trading) logic", SBP),
    ("item", "Failed order cooldown and retry logic", SBP),
    ("item", "Percent-of-account position sizing mode", SBP),
    ("item", "Fixed budget position sizing mode", SBP),
    ("item", "Fixed shares position sizing mode", SBP),
    ("item", "Fixed shares offset position sizing mode", SBP),

    ("category", "API Integration (Tradier)"),
    ("item", "Procedure to authenticate user (required for all APIs)", 0.5),
    ("item", "Procedure to fetch balances (needed for position sizing)", 0.5),
    ("item", "Procedure to fetch positions (prevent duplicate positions)", 0.5),
    ("item", "Procedure to fetch orders (monitor master account)", 0.5),
    ("item", "Procedure to send orders (including multi-leg)", 0.5),
    ("item", "Procedure to cancel orders", 0.5),
    ("item", "Procedure to fetch historical market data", P_ONLY),
    ("item", "Procedure to fetch real-time quotes", P_ONLY),

    # === PHASE 3: Programming Front-End ===
    ("phase", "Programming Front-End "),

    ("category", "Accounts page"),
    ("item", "Table displaying all follower accounts", 0.5),
    ("item", "Button to remove accounts on each row", 0.5),
    ("item", "Form to upload account API keys, store in database", 0.5),
    ("item", "Run validation checks on new accounts", 0.5),
    ("item", "Edit account alias and API key inline", P_ONLY),

    ("category", "Activity page"),
    ("item", "Table of logs showing bot actions", 0.5),

    ("category", "Balances page"),
    ("item", "Account equity cards (BOD, real-time, and change)", P_ONLY),
    ("item", "Line chart showing historical balances over time", P_ONLY),
    ("item", "Function to store current account balance each day", P_ONLY),

    ("category", "Chart page"),
    ("item", "Display OHLC candle chart", P_ONLY),
    ("item", "Toggle ticker and timeframe", P_ONLY),
    ("item", "Display OHLC data in table, exportable", P_ONLY),

    ("category", "Orders page"),
    ("item", "Table showing all orders by account", 0.5),
    ("item", "Cancel open order button", 0.5),
    ("item", "Toggle between database and broker order views", SBP),

    ("category", "PnL page"),
    ("item", "Retrieve all transactions from database", P_ONLY),
    ("item", "Calculate PnL for both short and long", P_ONLY),
    ("item", "Display results in a table", P_ONLY),
    ("item", "Ability to export table to Excel", P_ONLY),
    ("item", "Line chart summarizing the table", P_ONLY),

    ("category", "Positions page"),
    ("item", "Table showing all positions by account", 0.5),
    ("item", "Button to close position at market", 0.5),

    ("category", "Settings page"),
    ("item", "Master killswitch", 0.5),
    ("item", "Per-account position size multiplier input", 0.5),
    ("item", "Stale order timeout setting", 0.5),
    ("item", "Color mode (dark, light)", 0.5),
    ("item", "Password change", SBP),
    ("item", "Timezone selection", SBP),
    ("item", "Automation opt-in/opt-out toggle per user", SBP),
    ("item", "Order expiration threshold (minutes) input", SBP),
    ("item", "Trade direction (long/short) multi-select", SBP),
    ("item", "Trade reversal toggle", SBP),
    ("item", "Minimum position sizing toggle", SBP),
    ("item", "Position sizing mode selector with parameter inputs", SBP),

    ("category", "Trade page"),
    ("item", "Form to manually trade positions", P_ONLY),

    ("category", "Users page"),
    ("item", "Table showing all users", P_ONLY),
    ("item", "Button to delete user", P_ONLY),
    ("item", "Form to add new user", P_ONLY),

    ("category", "Admin and client views"),
    ("item", "Admin vs client filtered views on data pages", P_ONLY),

    ("category", "General branding"),
    ("item", "General branding (navbar logo, login logo, color scheme)", P_ONLY),

    # === PHASE 4: Finishing Details ===
    ("phase", "Finishing Details"),
    ("qa",),
    ("item", "Writing setup guide for things to do before project begins", 0.5),
    ("item", "Video conference setup tutorial after delivery", 0.5),
    ("item", "Expedited processing (none)", 0),
    ("item", "Consultation credit", -0.8),
]

# Validate all item descriptions are <= 72 chars
for entry in items:
    if entry[0] == "item":
        assert len(entry[1]) <= MAX_C_LEN, f"Item too long ({len(entry[1])}): {entry[1]}"
    elif entry[0] == "category":
        assert len(entry[1]) <= MAX_C_LEN, f"Category too long ({len(entry[1])}): {entry[1]}"

# --- Styles ---
center = Alignment(horizontal="center")
right = Alignment(horizontal="right")
bold_font = Font(bold=True)
header_fg = Color(theme=0, tint=-0.1499984740745262)
header_fill = PatternFill(patternType="solid", fgColor=header_fg)
title_fg = Color(theme=9, tint=0.5999938962981048)
title_fill = PatternFill(patternType="solid", fgColor=title_fg)
calc_font = Font(bold=True, color="FA7D00")
calc_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
note_font = Font(color="9C5700")
note_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
category_font = Font(bold=True, underline="single")
dollar_fmt = '"$"#,##0_);[Red]\\("$"#,##0\\)'
dollar_acct_fmt = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'
removed_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
added_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def get_hours(hours_val, tier_name):
    if isinstance(hours_val, dict):
        return hours_val.get(tier_name, 0)
    return hours_val


def build_sheet(ws, tier_name):
    is_base = (tier_name == BASE_TIER)
    ws.column_dimensions['B'].width = 25.73
    ws.column_dimensions['C'].width = 58.45
    ws.column_dimensions['D'].width = 10.63
    ws.column_dimensions['E'].width = 11.0
    ws.column_dimensions['F'].width = 10.54

    # Row 1: Title
    ws.merge_cells('B1:C1')
    ws['B1'] = "Cost Estimate"
    ws['B1'].fill = title_fill
    ws['B1'].alignment = center
    ws['C1'].fill = title_fill
    ws['E1'] = "Hourly Rate"
    ws['E1'].font = bold_font
    ws['E1'].fill = header_fill
    ws['F1'] = HOURLY_RATE
    ws['F1'].font = note_font
    ws['F1'].fill = note_fill
    ws['F1'].alignment = center

    # Row 3: Headers
    for col, header in [('B', 'Phase'), ('C', 'Item'), ('D', 'Hours'), ('E', 'Minutes'), ('F', 'Cost')]:
        cell = ws[f'{col}3']
        cell.value = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center

    ws.freeze_panes = 'A4'

    row = 4
    first_data_row = 4
    subtotal = 0

    for entry in items:
        if entry[0] == "phase":
            ws[f'B{row}'] = entry[1]
            row += 1

        elif entry[0] == "category":
            name = entry[1]
            ws[f'C{row}'] = name
            ws[f'C{row}'].font = category_font
            row += 1

        elif entry[0] == "item":
            desc, hours_val = entry[1], entry[2]
            hours = get_hours(hours_val, tier_name)
            ws[f'C{row}'] = desc
            ws[f'D{row}'] = hours
            ws[f'D{row}'].alignment = center
            ws[f'E{row}'] = f'=D{row}*60'
            ws[f'E{row}'].font = calc_font
            ws[f'E{row}'].fill = calc_fill
            ws[f'E{row}'].alignment = center
            ws[f'F{row}'] = f'=D{row}*F$1'
            ws[f'F{row}'].font = calc_font
            ws[f'F{row}'].fill = calc_fill
            ws[f'F{row}'].alignment = center
            ws[f'F{row}'].number_format = dollar_fmt
            if not is_base and isinstance(hours_val, dict):
                base_hours = get_hours(hours_val, BASE_TIER)
                if base_hours > 0 and hours == 0:
                    ws[f'C{row}'].fill = removed_fill
                    ws[f'D{row}'].fill = removed_fill
                elif base_hours == 0 and hours > 0:
                    ws[f'C{row}'].fill = added_fill
            subtotal += hours
            row += 1

        elif entry[0] == "qa":
            qa_text = "Final testing / debugging / trial runs / revisions / QA (10%)"
            ws[f'C{row}'] = qa_text
            ws[f'D{row}'] = f'=SUM(D{first_data_row}:D{row-1})*0.1'
            ws[f'D{row}'].font = calc_font
            ws[f'D{row}'].fill = calc_fill
            ws[f'D{row}'].alignment = center
            ws[f'E{row}'] = f'=D{row}*60'
            ws[f'E{row}'].font = calc_font
            ws[f'E{row}'].fill = calc_fill
            ws[f'E{row}'].alignment = center
            ws[f'F{row}'] = f'=D{row}*F$1'
            ws[f'F{row}'].font = calc_font
            ws[f'F{row}'].fill = calc_fill
            ws[f'F{row}'].alignment = center
            ws[f'F{row}'].number_format = dollar_fmt
            row += 1

    last_data_row = row - 1

    # Blank row
    row += 1

    # Totals row
    totals_row = row
    ws[f'C{row}'] = "Totals"
    ws[f'C{row}'].font = bold_font
    ws[f'C{row}'].fill = header_fill
    ws[f'C{row}'].alignment = right
    ws[f'D{row}'] = f'=SUM(D{first_data_row}:D{last_data_row})'
    ws[f'D{row}'].font = calc_font
    ws[f'D{row}'].fill = calc_fill
    ws[f'D{row}'].alignment = center
    ws[f'E{row}'] = f'=SUM(E{first_data_row}:E{last_data_row})'
    ws[f'E{row}'].font = calc_font
    ws[f'E{row}'].fill = calc_fill
    ws[f'E{row}'].alignment = center
    ws[f'F{row}'] = f'=SUM(F{first_data_row}:F{last_data_row})'
    ws[f'F{row}'].font = calc_font
    ws[f'F{row}'].fill = calc_fill
    ws[f'F{row}'].alignment = center
    ws[f'F{row}'].number_format = dollar_acct_fmt
    row += 1

    # Pace row
    pace_row = row
    ws[f'C{row}'] = "Average billable hours worked per day"
    ws[f'C{row}'].font = bold_font
    ws[f'C{row}'].fill = header_fill
    ws[f'C{row}'].alignment = right
    ws[f'D{row}'] = PACE
    ws[f'D{row}'].font = note_font
    ws[f'D{row}'].fill = note_fill
    ws[f'D{row}'].alignment = center
    ws[f'E{row}'] = "Start Date"
    ws[f'E{row}'].font = bold_font
    ws[f'E{row}'].fill = header_fill
    ws[f'E{row}'].alignment = right
    ws[f'F{row}'] = date.today()
    ws[f'F{row}'].number_format = 'YYYY-MM-DD'
    ws[f'F{row}'].font = note_font
    ws[f'F{row}'].fill = note_fill
    row += 1

    # Days row
    days_row = row
    ws[f'C{row}'] = "Days to complete"
    ws[f'C{row}'].font = bold_font
    ws[f'C{row}'].fill = header_fill
    ws[f'C{row}'].alignment = right
    ws[f'D{row}'] = f'=ROUNDUP(D{totals_row}/D{pace_row},0)'
    ws[f'D{row}'].font = calc_font
    ws[f'D{row}'].fill = calc_fill
    ws[f'D{row}'].alignment = center
    ws[f'E{row}'] = "End Date"
    ws[f'E{row}'].font = bold_font
    ws[f'E{row}'].fill = header_fill
    ws[f'E{row}'].alignment = right
    ws[f'F{row}'] = f'=F{pace_row}+D{days_row}'
    ws[f'F{row}'].number_format = 'YYYY-MM-DD'
    ws[f'F{row}'].font = calc_font
    ws[f'F{row}'].fill = calc_fill
    row += 1

    # Blank row
    row += 1

    # Notes
    ws[f'B{row}'] = "Notes"
    ws[f'B{row}'].font = bold_font
    ws[f'B{row}'].fill = header_fill
    row += 2

    notes = [
        "This estimate will expire 30 days after initial delivery",
        "Billable hours are usually less than total hours spent working on the project, so I will likely exceed the quoted hours above, but won't charge more than that",
        "I am currently working with 4-5 other clients, plus I have other responsibilites too, so I expect I can devote about 1-2 hours per day to this project on average",
        "Any changes to the project scope after the estimate has been provided may result in an updated estimate or additional charges.",
        "All software is provided on an as-is basis; use at your own risk.",
        "The estimate does not include costs for third-party services, software licenses, or external tools unless explicitly stated."
    ]
    for note in notes:
        ws[f'B{row}'] = note
        row += 1

    return subtotal


# --- Build workbook ---
wb = openpyxl.Workbook()

for i, tier_name in enumerate(TIERS):
    if i == 0:
        ws = wb.active
        ws.title = tier_name
    else:
        ws = wb.create_sheet(tier_name)
    subtotal = build_sheet(ws, tier_name)

    qa_hours = subtotal * 0.1
    total_hours = subtotal + qa_hours
    total_cost = total_hours * HOURLY_RATE
    days_to_complete = math.ceil(total_hours / PACE)
    print(f"[{tier_name}] Hours: {total_hours:.2f}, Cost: ${total_cost:,.0f}, Days: {days_to_complete}")

# Set Base sheet as active
base_index = TIERS.index(BASE_TIER)
wb.active = base_index

wb.save(output_path)
print(f"\nXLSX saved to {output_path}")

# ============================================================
# PDF Generation
# ============================================================
from fpdf import FPDF

# Pre-calculate tier totals and diffs
tier_totals = {}
for tn in TIERS:
    sub = sum(get_hours(e[2], tn) for e in items if e[0] == "item")
    qa = sub * 0.1
    total = sub + qa
    tier_totals[tn] = {'hours': total, 'cost': total * HOURLY_RATE,
                       'days': math.ceil(total / PACE)}

tier_diffs = {}
for tn in TIERS:
    if tn == BASE_TIER:
        continue
    removed, added = [], []
    for e in items:
        if e[0] == "item" and isinstance(e[2], dict):
            bh, th = get_hours(e[2], BASE_TIER), get_hours(e[2], tn)
            if bh > 0 and th == 0: removed.append(e[1])
            elif bh == 0 and th > 0: added.append(e[1])
    tier_diffs[tn] = {'removed': removed, 'added': added}

LOGO_PATH = r"C:\Users\twpot\Documents\Tyler's Stuff\Tyler's Documents\TNT Trading Partners\Cloud\tnttrading\static\img\upwork_logo.png"
PROFILE_NAME = "Tyler Potts"
PROFILE_URL = "upwork.com/freelancers/robotraderguy"

# BMP-only symbols
SYM_X      = '\u274c'
SYM_CHECK  = '\u2705'
SYM_WARN   = '\u26a0'
SYM_PENCIL = '\u270f'
SYM_GEAR   = '\u2699'
SYM_SCALES = '\u2696'
SYM_MEMO   = '\u2712'
SYM_STAR   = '\u2b50'
SYM_DASH   = '\u2014'
SYM_BULLET = '\u2022'

TIER_STARS = {"Simple": SYM_STAR, "Base": SYM_STAR * 2, "Premium": SYM_STAR * 3}


class EstimatePDF(FPDF):
    def header(self):
        self.image(LOGO_PATH, x=10, y=7, h=7)
        self.set_xy(19, 7)
        self.set_font('Segoe', 'B', 9)
        self.set_text_color(0, 30, 0)
        self.cell(0, 4, PROFILE_NAME)
        self.set_xy(19, 11)
        self.set_font('Segoe', '', 7)
        self.set_text_color(100, 100, 100)
        self.cell(0, 4, PROFILE_URL)
        self.set_draw_color(20, 168, 0)
        self.set_line_width(0.5)
        self.line(10, 16, self.w - 10, 16)
        self.set_text_color(0, 0, 0)
        self.set_draw_color(0, 0, 0)
        self.set_line_width(0.2)
        self.set_y(19)

    def footer(self):
        self.set_y(-15)
        self.set_font('Segoe', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', align='C')

    def emoji_prefix(self, emoji, text, h, font_style='B', font_size=13, align=''):
        self.set_font('SegoeEmoji', '', font_size)
        emoji_w = self.get_string_width(emoji)
        if emoji_w < 1:
            emoji_w = font_size * 0.5 * len(emoji)
        self.set_font('Segoe', font_style, font_size)
        text_w = self.get_string_width(text)
        total = emoji_w + text_w

        if align == 'C':
            self.set_x((self.w - total) / 2)

        self.set_font('SegoeEmoji', '', font_size)
        self.cell(emoji_w, h, emoji)
        self.set_font('Segoe', font_style, font_size)
        self.cell(0 if align != 'C' else text_w, h, text)


pdf = EstimatePDF()
pdf.alias_nb_pages()
pdf.set_auto_page_break(auto=True, margin=20)

pdf.add_font('Segoe', '', r'C:\Windows\Fonts\segoeui.ttf')
pdf.add_font('Segoe', 'B', r'C:\Windows\Fonts\segoeuib.ttf')
pdf.add_font('Segoe', 'I', r'C:\Windows\Fonts\segoeuii.ttf')
pdf.add_font('Segoe', 'BI', r'C:\Windows\Fonts\segoeuiz.ttf')
pdf.add_font('SegoeEmoji', '', r'C:\Windows\Fonts\seguiemj.ttf')

W = pdf.epw

# === Summary page ===
pdf.add_page()
pdf.emoji_prefix(SYM_PENCIL, ' Cost Estimate', 14, font_style='B', font_size=20, align='C')
pdf.ln(18)

pdf.emoji_prefix(SYM_GEAR, ' Project Description', 10, font_style='B', font_size=13)
pdf.ln(10)
pdf.set_font('Segoe', '', 10)
project_desc = (
    "A trade copier bot for Tradier that monitors a master account for new orders "
    "(including multi-leg options spreads like credit spreads on SPX and NDX) and "
    "automatically forwards them to up to 6 follower accounts. Each follower has a "
    "configurable position size multiplier (1x, 2x, 3x, etc.). The bot includes "
    "duplicate order prevention via database tracking, stale order auto-cancellation, "
    "and market hours awareness including early close detection."
)
pdf.multi_cell(0, 6, project_desc)
pdf.ln(10)

# Comparison table
pdf.emoji_prefix(SYM_SCALES, ' Tier Comparison', 10, font_style='B', font_size=13)
pdf.ln(12)

cw = [50, 35, 40, 50]
pdf.set_fill_color(70, 70, 70)
pdf.set_text_color(255, 255, 255)
pdf.set_font('Segoe', 'B', 11)
for h_text, w in zip(['Tier', 'Hours', 'Cost', 'Days to Complete'], cw):
    pdf.cell(w, 9, h_text, border=1, fill=True, align='C')
pdf.ln()
pdf.set_text_color(0, 0, 0)

for tn in TIERS:
    t = tier_totals[tn]
    is_b = tn == BASE_TIER
    pdf.set_font('Segoe', 'B' if is_b else '', 11)
    label = f'{tn} (recommended)' if is_b else tn
    pdf.cell(cw[0], 9, label, border=1, align='C')
    pdf.cell(cw[1], 9, f"{t['hours']:.1f}", border=1, align='C')
    pdf.cell(cw[2], 9, f"${t['cost']:,.0f}", border=1, align='C')
    pdf.cell(cw[3], 9, str(t['days']), border=1, align='C')
    pdf.ln()

pdf.ln(10)

# Feature diffs
for tn in TIERS:
    if tn == BASE_TIER:
        continue
    d = tier_diffs[tn]
    if not d['removed'] and not d['added']:
        continue
    pdf.set_font('Segoe', 'B', 12)
    pdf.cell(0, 10, f'{tn} Tier {SYM_DASH} Changes from Base')
    pdf.ln(10)
    if d['removed']:
        for item in d['removed']:
            pdf.set_x(pdf.l_margin + 8)
            pdf.set_text_color(180, 0, 0)
            pdf.set_font('SegoeEmoji', '', 10)
            pdf.cell(6, 7, SYM_X)
            pdf.set_font('Segoe', '', 10)
            pdf.cell(0, 7, item)
            pdf.ln()
        pdf.set_text_color(0, 0, 0)
        pdf.ln(2)
    if d['added']:
        for item in d['added']:
            pdf.set_x(pdf.l_margin + 8)
            pdf.set_text_color(0, 130, 0)
            pdf.set_font('SegoeEmoji', '', 10)
            pdf.cell(6, 7, SYM_CHECK)
            pdf.set_font('Segoe', '', 10)
            pdf.cell(0, 7, item)
            pdf.ln()
        pdf.set_text_color(0, 0, 0)
    pdf.ln(5)

# === Per-tier detail pages ===
for tn in TIERS:
    pdf.add_page()
    t = tier_totals[tn]

    stars = TIER_STARS.get(tn, '')
    pdf.emoji_prefix(stars, f' {tn} Tier', 10, font_style='B', font_size=16, align='C')
    pdf.ln(8)
    pdf.set_font('Segoe', '', 9)
    pdf.cell(0, 5, f'Rate: ${HOURLY_RATE}/hr  |  Pace: {PACE} hrs/day', align='C')
    pdf.ln(5)

    iw = W - 55
    hw = 25
    cstw = 30
    RH = 6

    # Table header
    pdf.set_fill_color(70, 70, 70)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Segoe', 'B', 9)
    pdf.cell(iw, 7, 'Item', border=1, fill=True)
    pdf.cell(hw, 7, 'Hours', border=1, fill=True, align='C')
    pdf.cell(cstw, 7, 'Cost', border=1, fill=True, align='C')
    pdf.ln()
    pdf.set_text_color(0, 0, 0)

    row_i = 0
    for entry in items:
        if entry[0] == "phase":
            pdf.set_font('Segoe', 'B', 9)
            pdf.set_fill_color(220, 220, 220)
            pdf.cell(W, 7, entry[1].strip(), border='LR', fill=True)
            pdf.ln()

        elif entry[0] == "category":
            pdf.set_font('Segoe', 'BI', 8)
            pdf.cell(iw, RH, f'  {entry[1]}', border='L')
            pdf.cell(hw, RH, '', border=0)
            pdf.cell(cstw, RH, '', border='R')
            pdf.ln()

        elif entry[0] == "item":
            hours = get_hours(entry[2], tn)
            cost = hours * HOURLY_RATE
            is_removed = is_added = False
            if tn != BASE_TIER and isinstance(entry[2], dict):
                bh = get_hours(entry[2], BASE_TIER)
                if bh > 0 and hours == 0: is_removed = True
                elif bh == 0 and hours > 0: is_added = True

            alt = row_i % 2 == 0
            if alt:
                pdf.set_fill_color(248, 248, 248)

            if is_removed: pdf.set_text_color(180, 0, 0)
            elif is_added: pdf.set_text_color(0, 130, 0)
            elif hours == 0: pdf.set_text_color(160, 160, 160)

            pdf.set_font('Segoe', '', 8)
            pdf.cell(iw, RH, f'  {entry[1]}', border='L', fill=alt)
            pdf.cell(hw, RH, str(hours), border=0, fill=alt, align='C')
            pdf.cell(cstw, RH, f'${cost:,.0f}', border='R', fill=alt, align='C')
            pdf.ln()
            pdf.set_text_color(0, 0, 0)
            row_i += 1

        elif entry[0] == "qa":
            qa_hrs = sum(get_hours(e[2], tn) for e in items if e[0] == "item") * 0.1
            qa_cost = qa_hrs * HOURLY_RATE
            alt = row_i % 2 == 0
            if alt:
                pdf.set_fill_color(248, 248, 248)
            pdf.set_font('Segoe', '', 8)
            pdf.cell(iw, RH, '  Final testing / debugging / revisions / QA (10%)', border='L', fill=alt)
            pdf.cell(hw, RH, f'{qa_hrs:.1f}', border=0, fill=alt, align='C')
            pdf.cell(cstw, RH, f'${qa_cost:,.0f}', border='R', fill=alt, align='C')
            pdf.ln()
            row_i += 1

    # Totals row
    if pdf.get_y() > pdf.h - pdf.b_margin - 18:
        pdf.add_page()
    pdf.set_font('Segoe', 'B', 10)
    pdf.set_fill_color(70, 70, 70)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(iw, 9, 'Totals', border=1, fill=True, align='R')
    pdf.cell(hw, 9, f"{t['hours']:.1f}", border=1, fill=True, align='C')
    pdf.cell(cstw, 9, f"${t['cost']:,.0f}", border=1, fill=True, align='C')
    pdf.ln()
    pdf.set_text_color(0, 0, 0)

    # Timeline
    pdf.ln(3)
    pdf.set_font('Segoe', '', 9)
    start = date.today()
    end = start + timedelta(days=t['days'])
    pdf.cell(0, 6, f"Start: {start.strftime('%B %d, %Y')}  |  Est. completion: {end.strftime('%B %d, %Y')}", align='C')
    pdf.ln()

# === Notes page ===
pdf.add_page()
pdf.emoji_prefix(SYM_MEMO, ' Notes', 10, font_style='B', font_size=14)
pdf.ln(15)
pdf.set_font('Segoe', '', 10)
pdf_notes = [
    "This estimate will expire 30 days after initial delivery.",
    "Billable hours are usually less than total hours spent working on the project, so I will likely exceed the quoted hours above, but won't charge more than that.",
    "I am currently working with 4-5 other clients, plus I have other responsibilities too, so I expect I can devote about 1-2 hours per day to this project on average.",
    "Any changes to the project scope after the estimate has been provided may result in an updated estimate or additional charges.",
    "All software is provided on an as-is basis; use at your own risk.",
    "The estimate does not include costs for third-party services, software licenses, or external tools unless explicitly stated."
]
for n in pdf_notes:
    pdf.set_font('Segoe', '', 10)
    pdf.multi_cell(0, 6, f'{SYM_BULLET} {n}')
    pdf.ln(3)

# Disclaimer
pdf.ln(10)
pdf.set_draw_color(200, 200, 200)
pdf.line(10, pdf.get_y(), pdf.w - 10, pdf.get_y())
pdf.ln(5)
pdf.set_text_color(120, 120, 120)
pdf.set_font('SegoeEmoji', '', 7)
pdf.cell(4, 4, SYM_WARN)
pdf.set_font('Segoe', 'I', 7)
disclaimer_text = (
    "DISCLAIMER: Tyler Potts is an independent contractor on the Upwork "
    "platform and is not an employee, agent, or representative of Upwork Inc. This "
    "estimate is provided independently and does not constitute an endorsement, "
    "guarantee, or recommendation by Upwork Inc. All work is performed under a direct "
    "agreement between the freelancer and the client. The Upwork logo is a trademark "
    "of Upwork Inc. and is used here solely to identify the platform through which "
    "the freelancer offers services."
)
pdf.multi_cell(0, 4, disclaimer_text)
pdf.set_text_color(0, 0, 0)

pdf_path = output_path.replace('.xlsx', '.pdf')
pdf.output(pdf_path)
print(f"PDF saved to {pdf_path}")

# ============================================================
# Update ConsultationStats.xlsx
# ============================================================
stats_path = f"{CONSULTATIONS_DIR}\\ConsultationStats.xlsx"
try:
    swb = openpyxl.load_workbook(stats_path)
    sws = swb['Rates']

    # Use Base tier for stats
    base_total = tier_totals[BASE_TIER]
    target_row = CONSULTATION_NUMBER + 1  # Row 2 = consultation 1

    # Check if row already exists
    existing = sws[f'A{target_row}'].value
    if existing is not None and existing != CONSULTATION_NUMBER:
        print(f"WARNING: Row {target_row} already has consultation {existing}, skipping stats update")
    else:
        sws[f'A{target_row}'] = CONSULTATION_NUMBER
        sws[f'B{target_row}'] = date.today()
        sws[f'C{target_row}'] = HOURLY_RATE
        sws[f'D{target_row}'] = base_total['hours']
        sws[f'E{target_row}'] = base_total['cost']
        sws[f'F{target_row}'] = PACE
        sws[f'G{target_row}'] = base_total['days']
        swb.save(stats_path)
        print(f"ConsultationStats updated: row {target_row}")
except Exception as e:
    print(f"Stats update error (non-fatal): {e}")

print("\nDone!")
